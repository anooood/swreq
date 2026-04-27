"""
hierarchy_sheet.py
------------------
Adds a single hierarchy sheet to the workbook:

  Jama_Hierarchy   — plain text, indentation-only

The previous color-coded Jama_Hierarchy_Color sheet has been removed at
the user's request — only the plain hierarchy is needed for Jama import.

Columns:
  Level | Name | Description | VerificationMethod | RequirementType | EngineeringDiscipline
"""

from __future__ import annotations

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Shared constants
# ---------------------------------------------------------------------------

COLS = [
    "Level",
    "Name",
    "Description",
    "VerificationMethod",
    "RequirementType",
    "EngineeringDiscipline",
]
COL_WIDTHS = {1: 8, 2: 55, 3: 70, 4: 22, 5: 22, 6: 24}

PLAIN_FONT  = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=10, bold=True)
CENTER      = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Plain row helpers
# ---------------------------------------------------------------------------

def _plain_row(ws, row_idx: int, values: list[str], height: int = 15, indent: int = 0):
    ws.row_dimensions[row_idx].height = height
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.font = PLAIN_FONT
        # Apply Excel native indent to the Name column (col 2) only
        c.alignment = Alignment(
            horizontal="left",
            wrap_text=True,
            vertical="top",
            indent=(indent if col == 2 else 0),
        )


def _plain_section(ws, row_idx: int, level: int, name: str):
    _plain_row(ws, row_idx, [str(level), name, "", "", "", ""], indent=level)


def _plain_req(ws, row_idx: int, level: int, hlr):
    _plain_row(
        ws,
        row_idx,
        [
            str(level),
            hlr.name,
            hlr.description,
            hlr.verification_method,
            hlr.requirement_type,
            hlr.engineering_discipline,
        ],
        height=30,
        indent=level,
    )


# ---------------------------------------------------------------------------
# Header writer
# ---------------------------------------------------------------------------

def _write_header(ws):
    for j, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font      = HEADER_FONT
        c.alignment = CENTER
    ws.row_dimensions[1].height = 18


# ---------------------------------------------------------------------------
# Tree builder
# ---------------------------------------------------------------------------

def _build_tree(ws, result, module_name: str):
    row = 2
    hlrs_by_func: dict[str, list] = defaultdict(list)
    for hlr in result.hlrs:
        hlrs_by_func[hlr.function_name].append(hlr)

    _plain_section(ws, row, 0, module_name);                row += 1
    _plain_section(ws, row, 1, "Scope");                    row += 1
    _plain_section(ws, row, 1, "System Overview");          row += 1
    _plain_section(ws, row, 1, "Requirements");             row += 1
    _plain_section(ws, row, 2, "Required States and Modes"); row += 1
    _plain_section(ws, row, 2, "Capability Requirements");  row += 1

    for group in result._groups:
        func_hlrs = hlrs_by_func.get(group.function_name, [])
        if not func_hlrs:
            continue
        _plain_section(ws, row, 3, group.function_name);    row += 1
        for hlr in func_hlrs:
            _plain_req(ws, row, 4, hlr);                    row += 1

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_hierarchy_sheet(
    wb: Workbook,
    result,
    module_name: str = "Software Requirements for Module",
):
    """
    Add a single "Jama_Hierarchy" sheet to wb (plain, indentation-only).
    Returns the worksheet so callers can chain styling if needed.
    """
    ws = wb.create_sheet("Jama_Hierarchy")
    ws.sheet_view.showGridLines = False
    _write_header(ws)
    _build_tree(ws, result, module_name)
    return ws
