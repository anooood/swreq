import streamlit as st
import time
import requests
import glob
from pathlib import Path
import pandas as pd
import regex as re
import torch
import json
import ast as asast
from pre_processing import (
    ContextExtractor,
    pre_processing,
    generate_functions,
    extract_header_comments,
    clean_header_comments,
    get_header_function_descriptions,
    get_identifier_mapping,
    get_comment_positions,
    add_comments,
    extract_function_context,
)
from openpyxl import load_workbook
from prompt_templates_new import name_prompt_template, reqs_prompt_template, second_prompt, identifier_rewrite_prompt
from word_output import generate_doc_from_cleaned_and_df

from pycparser import c_generator, parse_file

# ─────────────────────────────
# Streamlit config
# ─────────────────────────────
st.set_page_config(layout="wide")
st.title("MCP Application Requirements Generator")

# ─────────────────────────────
# Utilities
# ─────────────────────────────
def strip_markdown(text: str) -> str:
    if not text:
        return ""
    text = text.replace('"', '')
    text = re.sub(r'```.*?```', '', text, flags=re.DOTALL)
    text = re.sub(r'`([^`]*)`', r'\1', text)
    text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)
    text = re.sub(r'\*(.*?)\*', r'\1', text)
    text = re.sub(r'#+\s*', '', text)
    text = re.sub(r'[-•]\s+', '', text)
    text = re.sub(r'\n{2,}', '\n', text)
    return text.strip()


def auto_height(text, min_lines=4, max_lines=30, line_px=22):
    lines = max(text.count("\n") + 1, min_lines)
    lines = min(lines, max_lines)
    return (lines * line_px) + 100


def format_globals(globals_dict):
    if not globals_dict:
        return "_No global variables referenced._"
    lines = ["**Global variables used:**"]
    for k, v in globals_dict.items():
        lines.append(f"- `{k}` = `{v}`")
    return "\n".join(lines)

def on_module_change():
    st.session_state["requirements"].clear()
    st.session_state["functions"] = []
    st.session_state["cleaned_headers"] = None
    st.session_state["generation_started"] = False
    st.session_state["generation_completed"] = False

    # 🔥 clear editor widgets
    for k in list(st.session_state.keys()):
        if k.startswith("title_") or k.startswith("body_"):
            del st.session_state[k]
    
def append_df_to_excel(file_path, df, sheet_name="Sheet1"):
    file_path = Path(file_path)

    if not file_path.exists():
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
        return

    with pd.ExcelWriter(
        file_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        if sheet_name in writer.book.sheetnames:
            start_row = writer.book[sheet_name].max_row
        else:
            start_row = 0

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False,
            header=start_row == 0
        )

IDENTIFIER = re.compile(r'[A-Za-z_][A-Za-z0-9_.]*')

def replace_exact_identifiers(text, mapping):
    def replacer(match):
        token = match.group(0)
        return mapping.get(token, token)

    return IDENTIFIER.sub(replacer, text)

def split_merged_words_keep_all_caps(token):
    # Leave pure ALL-CAPS words untouched
    if token.isupper():
        return token

    # Split camelCase / PascalCase
    token = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', token)

    return token.lower()

def remove_consecutive_duplicates(text):
    return re.sub(r'\b(\w+)(?:\s+\1\b)+', r'\1', text, flags=re.IGNORECASE)
    
# ─────────────────────────────
# Persistent state (single source of truth)
# ─────────────────────────────
if "generation_started" not in st.session_state:
    st.session_state["generation_started"] = False

if "generation_completed" not in st.session_state:
    st.session_state["generation_completed"] = False

if "functions" not in st.session_state:
    st.session_state["functions"] = []

if "requirements" not in st.session_state:
    st.session_state["requirements"] = {}
    
# ─────────────────────────────
# Module selection
# ─────────────────────────────
Git_folder = "P3_MCP_Application/cmake-src/src"

modules = [Path(f).name for f in glob.glob(Git_folder + "/*.c")]
modules_names = [x.name.replace(".c","") for x in sorted(Path(Git_folder).glob("*.c"))]

selected_module = st.selectbox(
    "Select a system/module:",
    modules,
    on_change=on_module_change
)

with open(Path(Git_folder) / selected_module, "r", encoding="utf-8", errors="ignore") as f:
    source_code = f.read()


# ─────────────────────────────
# Generate requirements
# ─────────────────────────────
if st.button("Generate Requirements"):
    st.session_state["requirements"] = {}
    st.session_state["functions"] = []
    st.session_state["generation_started"] = True
    st.session_state["generation_completed"] = False
    st.session_state["variables_leaked_all"] = []
    st.session_state["globals_missing_all"] = []
    st.session_state["inference_times"] = []

functions = st.session_state.get("functions", [])
generator = st.session_state.get("ast_generator")
extractor = st.session_state.get("extractor")

if st.session_state["generation_started"] and not st.session_state["generation_completed"]:

    with st.spinner("Generating requirements..."):

        output_file, input_file = pre_processing(selected_module)
        functions = generate_functions(output_file, input_file)
        st.session_state["functions"] = functions

        cleaned_headers = clean_header_comments(
            extract_header_comments(input_file)
        )
        st.session_state["cleaned_headers"] = cleaned_headers

        ast = parse_file(output_file)
        extractor = ContextExtractor()
        extractor.visit(ast)
        generator = c_generator.CGenerator()
        function_context = extract_function_context(extractor)

        st.subheader(f"Requirements for **{selected_module}**")
        progress = st.progress(0)

        for i, fn in enumerate(functions):
            # Each function has its own placeholder
            placeholder = st.empty()

            # --- Generate requirement if not yet generated ---
            if i not in st.session_state["requirements"]:
                start_time = time.time()

                orig_src = fn["content"]
                func_node = extractor.function_defs.get(fn['name'])
                emb_src = generator.visit(func_node)
                mapping = get_identifier_mapping(orig_src, emb_src)

                name_prompt = name_prompt_template.format(
                    header_comments=st.session_state["cleaned_headers"],
                    function_name=fn["name"]
                )
                reqs_prompt = reqs_prompt_template.format(
                    C_FUNCTION=fn["content"],
                    GLOBAL_VARIABLES=mapping
                )

                response = requests.post(
                    "http://127.0.0.1:8000/generate",
                    json={"prompts": [name_prompt, reqs_prompt]},
                    timeout=300
                )

                if response.status_code != 200:
                    placeholder.error(f"Model failed for {fn['name']}")
                    continue

                data = response.json()
                torch.cuda.empty_cache()

                fixed_req = data["requirements"][1]
                for k, v in mapping.items():
                    cleaned_v = re.sub(r"\s*\([^)]*\)", "", v)
                    fixed_req = fixed_req.replace(k, cleaned_v)

                ## CHECK FOR LEAKED SYNTAX AND MISSING GLOBALS
                variables_leaked = []
                for var in function_context[functions[i]['name']]['globals_used']:
                    if var in modules_names:
                        continue
                    if var in fixed_req:
                        variables_leaked.append(var)
                for local in function_context[functions[i]['name']]['locals_declared']:
                    if local in modules_names:
                        continue
                    if local in fixed_req:
                        variables_leaked.append(local)
                for func in function_context[functions[i]['name']]['functions_called']:
                    if func in fixed_req:
                        variables_leaked.append(func)
                for struct in function_context[functions[i]['name']]['used_structures']:
                    if struct in modules_names:
                        continue
                    if struct in fixed_req:
                        variables_leaked.append(struct)
                if functions[i]['name'] in fixed_req:
                    variables_leaked.append(functions[i]['name'])

                if len(variables_leaked) != 0:
                    ## SECOND PASS TO REMOVE LEAKED SYNTAX
                    post_process_prompt = identifier_rewrite_prompt.format(
                        REQUIREMENTS=fixed_req,
                        VARIABLES=variables_leaked
                    )
                    response_2 = requests.post(
                        "http://127.0.0.1:8000/generate",
                        json={"prompts": ["N/A", post_process_prompt]},
                        timeout=300
                    )
                    data_2 = response_2.json()
                    variables_leaked_fixed = asast.literal_eval(data_2["requirements"][1])
                    variables_mapping = dict(zip(variables_leaked, variables_leaked_fixed))

                    final_req = replace_exact_identifiers(fixed_req, variables_mapping)
                    final_req = remove_consecutive_duplicates(final_req)
                else:
                    final_req = fixed_req
                    
                end_time = time.time()
                response_time = end_time - start_time
                st.session_state["inference_times"].append(response_time)

                for var in function_context[functions[i]['name']]['globals_used']:
                    if var in modules_names:
                        continue
                    if var in final_req:
                        st.session_state["variables_leaked_all"].append(var)
                for local in function_context[functions[i]['name']]['locals_declared']:
                    if local in modules_names:
                        continue
                    if local in final_req:
                        st.session_state["variables_leaked_all"].append(local)
                for func in function_context[functions[i]['name']]['functions_called']:
                    if func in final_req:
                        st.session_state["variables_leaked_all"].append(func).append(func)
                for struct in function_context[functions[i]['name']]['used_structures']:
                    if struct in modules_names:
                        continue
                    if struct in final_req:
                        st.session_state["variables_leaked_all"].append(struct)
                if functions[i]['name'] in final_req:
                    st.session_state["variables_leaked_all"].append(functions[i]['name'])
                for k, v in mapping.items():
                    cleaned_v = re.sub(r"\s*\([^)]*\)", "", v)
                    cleaned_v = cleaned_v.replace(')','').strip()
                    if cleaned_v not in final_req:
                        st.session_state["globals_missing_all"].append(cleaned_v)

                st.session_state["requirements"][i] = {
                    "Function": strip_markdown(data["requirements"][0]),
                    "Requirements": final_req.strip(),
                    "ModelResponseTimeSec": response_time,
                    "Applied": False,
                    "Globals": mapping,
                }

            # --- Replace placeholder with editable form immediately ---
            req = st.session_state["requirements"][i]
            col_req, col_code = placeholder.columns([1, 1])

            with col_req:
                form_key = f"{selected_module}_form_{i}"
                title_key = f"{selected_module}_title_{i}"
                body_key = f"{selected_module}_body_{i}"

                with st.form(form_key, clear_on_submit=False):
                    title = st.text_input(
                        "Requirement Title",
                        value=req["Function"],
                        key=title_key
                    )
                    body = st.text_area(
                        "Requirement Text",
                        value=req["Requirements"],
                        key=body_key,
                        height=auto_height(req["Requirements"])
                    )

                    if st.form_submit_button("Apply changes"):
                        st.session_state["requirements"][i].update({
                            "Function": st.session_state[title_key],
                            "Requirements": st.session_state[body_key],
                            "Applied": True,
                        })
                        st.success("Changes applied")

                st.markdown(format_globals(req["Globals"]))

                st.markdown(f"**Response Time:** {round(req['ModelResponseTimeSec'],2)} s")

            with col_code:
                with st.expander("Code Reference", expanded=True):
                    st.code(fn["content"], language="c")

            st.divider()
            progress.progress((i + 1) / len(functions))

# ─────────────────────────────
# Export (ALWAYS fresh)
# ─────────────────────────────
st.subheader("Export")

if st.session_state["requirements"]:
    st.markdown(f"**Total Inference Time:** {round(sum(st.session_state["inference_times"])/60,2)} min")
    df_latest = pd.DataFrame.from_dict(
        st.session_state["requirements"],
        orient="index"
    )

    if st.button("Prepare Word document"):
        row = {
            "Module": selected_module,
            "Leaked Syntax": json.dumps(
                list(set(st.session_state.get("variables_leaked_all", [])))
            ),
            "Missing Globals": json.dumps(
                list(set(st.session_state.get("globals_missing_all", [])))
            )
        }

        df_log = pd.DataFrame([row])
        append_df_to_excel("results.xlsx", df_log)

        cleaned_module_name = selected_module.replace(".c","")
        DOCX_PATH = Path(f"{cleaned_module_name}_module_reqs.docx")

        generate_doc_from_cleaned_and_df(
            df=df_latest,
            cleaned=st.session_state["cleaned_headers"],
            out_path=str(DOCX_PATH),
            debug=True
        )

        st.download_button(
            label="Download Module Spec (Word)",
            data=DOCX_PATH.read_bytes(),
            file_name=f"{cleaned_module_name}_module_reqs.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

    if st.button("Prepare CSV"):
        cleaned_module_name = selected_module.replace(".c", "")

        rows = []
        for _, row in df_latest.iterrows():
            function_name = row["Function"]
            requirements = row["Requirements"]
            rows.append({
                "Function Name": function_name,
                "Requirements": requirements,
            })
            # # Each requirement on its own line → split and explode
            # for req in row["Requirements"].splitlines():
            #     req = req.strip()
            #     if req:
            #         rows.append({
            #             "Function Name": function_name,
            #             "Requirements": req,
            #         })

        df_csv = pd.DataFrame(rows, columns=["Function Name", "Requirements"])
        csv_bytes = df_csv.to_csv(index=False).encode("utf-8")

        st.download_button(
            label="Download Requirements (CSV)",
            data=csv_bytes,
            file_name=f"{cleaned_module_name}_requirements.csv",
            mime="text/csv",
        )

else:
    st.info("No requirements available yet.")

# ─────────────────────────────
# Export (ALWAYS fresh)
# ─────────────────────────────
# st.subheader("Export")

# if st.session_state["requirements"]:
#     st.markdown(f"**Total Inference Time:** {round(sum(st.session_state["inference_times"])/60,2)} min")
#     df_latest = pd.DataFrame.from_dict(
#         st.session_state["requirements"],
#         orient="index"
#     )
#     print(df_latest)

#     if st.button("Prepare Word document"):
#         row = {
#             "Module": selected_module,
#             "Leaked Syntax": json.dumps(
#                 list(set(st.session_state.get("variables_leaked_all", [])))
#             ),
#             "Missing Globals": json.dumps(
#                 list(set(st.session_state.get("globals_missing_all", [])))
#             )
#         }

#         df_log = pd.DataFrame([row])
#         append_df_to_excel("results.xlsx", df_log)

#         cleaned_module_name = selected_module.replace(".c","")
#         DOCX_PATH = Path(f"{cleaned_module_name}_module_reqs.docx")

#         generate_doc_from_cleaned_and_df(
#             df=df_latest,
#             cleaned=st.session_state["cleaned_headers"],
#             out_path=str(DOCX_PATH),
#             debug=True
#         )

#         st.download_button(
#             label="Download Module Spec (Word)",
#             data=DOCX_PATH.read_bytes(),
#             file_name=f"{cleaned_module_name}_module_reqs.docx",
#             mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
#         )
# else:
#     st.info("No requirements available yet.")