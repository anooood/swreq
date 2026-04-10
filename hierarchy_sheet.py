"""
hierarchy_sheet.py
------------------
Adds TWO hierarchy sheets to the workbook:

  Jama_Hierarchy        — plain text, indentation-only (no colors)
  Jama_Hierarchy_Color  — full-row background color per level, as supported
                          by Jama's Excel color-based hierarchy import.

Color mapping (Jama convention — entire row background):
  Level 0  #1F3864  dark navy    — document root
  Level 1  #2E75B6  mid blue     — top-level sections
  Level 2  #4472C4  cornflower   — sub-sections
  Level 3  #BDD7EE  pale blue    — function headings
  Level 4  #FFFFFF  white        — individual requirements (leaf nodes)

Both sheets share the same columns:
  Level | Name | Description | VerificationMethod | RequirementType
"""

from __future__ import annotations

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Shared constants
# ---------------------------------------------------------------------------


COLS       = ["Level", "Name", "Description", "VerificationMethod", "RequirementType"]
COL_WIDTHS = {1: 8, 2: 55, 3: 70, 4: 22, 5: 22}

PLAIN_FONT  = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=10, bold=True)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")

# Level → (background hex, font color hex, bold, row height px)
LEVEL_STYLE = {
    0: ("1F3864", "FFFFFF", True,  20),   # dark navy,    white text
    1: ("2E75B6", "FFFFFF", True,  18),   # mid blue,     white text
    2: ("4472C4", "FFFFFF", True,  18),   # cornflower,   white text
    3: ("BDD7EE", "1F3864", True,  16),   # pale blue,    dark text
    4: ("FFFFFF", "000000", False, 30),   # white,        black text
}

THIN_BORDER = Border(
    left=Side(style="thin",   color="D9D9D9"),
    right=Side(style="thin",  color="D9D9D9"),
    top=Side(style="thin",    color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)



# ---------------------------------------------------------------------------
# Plain sheet helpers (no color)
# ---------------------------------------------------------------------------

def _plain_row(ws, row_idx: int, values: list[str], height: int = 15, indent: int = 0):
    ws.row_dimensions[row_idx].height = height
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.font      = PLAIN_FONT
        # Apply Excel native indent to the Name column (col 2) only
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top",
                                indent=(indent if col == 2 else 0))


def _plain_section(ws, row_idx: int, level: int, name: str):
    _plain_row(ws, row_idx, [str(level), name, "", "", ""], indent=level)


def _plain_req(ws, row_idx: int, level: int, hlr):
    _plain_row(ws, row_idx, [
        str(level),
        hlr.name,
        hlr.description,
        hlr.verification_method,
        hlr.requirement_type,
    ], height=30, indent=level)


# ---------------------------------------------------------------------------
# Color sheet helpers
# ---------------------------------------------------------------------------

def _color_row(ws, row_idx: int, level: int, values: list[str]):
    bg, fg, bold, height = LEVEL_STYLE[level]
    fill = PatternFill("solid", start_color=bg)
    font = Font(name="Arial", size=10, bold=bold, color=fg)
    ws.row_dimensions[row_idx].height = height
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.fill      = fill
        c.font      = font
        c.border    = THIN_BORDER
        c.alignment = WRAP


def _color_section(ws, row_idx: int, level: int, name: str):
    _color_row(ws, row_idx, level, [str(level), name, "", "", ""])


def _color_req(ws, row_idx: int, level: int, hlr):
    _color_row(ws, row_idx, level, [
        str(level),
        hlr.name,
        hlr.description,
        hlr.verification_method,
        hlr.requirement_type,
    ])


# ---------------------------------------------------------------------------
# Shared header writer
# ---------------------------------------------------------------------------

def _write_header(ws, colored: bool = False):
    for j, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font      = HEADER_FONT
        c.alignment = CENTER
        if colored:
            c.fill   = PatternFill("solid", start_color="1F3864")
            c.font   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.border = THIN_BORDER
    ws.row_dimensions[1].height = 18


# ---------------------------------------------------------------------------
# Shared tree builder — calls the appropriate row writers
# ---------------------------------------------------------------------------

def _build_tree(ws, result, module_name: str, section_fn, req_fn):
    row = 2
    hlrs_by_func: dict[str, list] = defaultdict(list)
    for hlr in result.hlrs:
        hlrs_by_func[hlr.function_name].append(hlr)

    section_fn(ws, row, 0, module_name);               row += 1
    section_fn(ws, row, 1, "Scope");                   row += 1
    section_fn(ws, row, 1, "System Overview");         row += 1
    section_fn(ws, row, 1, "Requirements");            row += 1
    section_fn(ws, row, 2, "Required States and Modes"); row += 1
    section_fn(ws, row, 2, "Capability Requirements"); row += 1

    for group in result._groups:
        func_hlrs = hlrs_by_func.get(group.function_name, [])
        if not func_hlrs:
            continue
        section_fn(ws, row, 3, group.function_name);   row += 1
        for hlr in func_hlrs:
            req_fn(ws, row, 4, hlr);                   row += 1

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_hierarchy_sheet(
    wb: Workbook,
    result,
    module_name: str = "Software Requirements for Module",
):
    """
    Adds both hierarchy sheets to wb:
      - Jama_Hierarchy        : plain, indentation only
      - Jama_Hierarchy_Color  : full-row background color per level
    """
    # Sheet 1 — plain
    ws_plain = wb.create_sheet("Jama_Hierarchy")
    ws_plain.sheet_view.showGridLines = False
    _write_header(ws_plain, colored=False)
    _build_tree(ws_plain, result, module_name, _plain_section, _plain_req)

    # Sheet 2 — colored
    ws_color = wb.create_sheet("Jama_Hierarchy_Color")
    ws_color.sheet_view.showGridLines = False
    _write_header(ws_color, colored=True)
    _build_tree(ws_color, result, module_name, _color_section, _color_req)

    return ws_plain, ws_color
