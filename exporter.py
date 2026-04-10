"""
exporter.py
-----------
Converts a MergeResult into a Jama-import-compatible Excel workbook.

Sheets:
  1. Jama_Requirements  — one HLR per row, exact Jama field names as headers
  2. Traceability       — one row per LLR, linked to its parent HLR
  3. Function_Summary   — LLR count and HLR count per function
  4. Jama_Hierarchy     — indented tree for Jama hierarchical import
"""

from __future__ import annotations

import io
from collections import defaultdict
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from merger import MergeResult

# ---------------------------------------------------------------------------
# Style palette
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="1F3864")   # dark navy
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FILL  = PatternFill("solid", start_color="2E75B6")   # mid blue
TITLE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=12)
BODY_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")   # very light blue
WARN_FILL   = PatternFill("solid", start_color="FFF2CC")   # pale amber
THIN        = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _banner(ws, text: str, ncols: int, row: int = 1):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font  = TITLE_FONT
    c.fill  = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def _headers(ws, cols: list[str], row: int = 2):
    for j, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.border    = THIN
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22


def _cell(ws, row: int, col: int, val: str, alt=False, warn=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = BODY_FONT
    c.border    = THIN
    c.alignment = WRAP
    if warn:
        c.fill = WARN_FILL
    elif alt:
        c.fill = ALT_FILL
    return c


def _widths(ws, mapping: dict[int, int]):
    for col, w in mapping.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------------------
# Description formatter
# ---------------------------------------------------------------------------

def _fmt_description(text: str) -> str:
    """
    Split a description on clause-separator dashes (' - ') so each clause
    appears on its own line in the Excel cell.

    Correctly preserves:
      - Negative numbers:        "offset of -40"
      - Arithmetic expressions:  "raw_value * 0.1 - 40"
      - Formula subtractions:    "(raw * 0.0625) - 273.15"
      - Numeric ranges:          "clamp to 0-100"

    Splits on:
      - Clause separators:       "validate - reject - log"
      - Mixed numeric/clause:    "scale by 0.5 - clamp to range"
    """
    import re as _re

    OPERATORS = set('+-*/=<>^%|&')

    result = []
    last = 0
    for m in _re.finditer(r'\s+-\s+', text):
        before_text = text[last:m.start()]
        after_text  = text[m.end():]
        after_char  = after_text[0] if after_text else ''

        # First token after the dash
        after_token = _re.split(r'\s+', after_text.strip())[0] if after_text.strip() else ''
        after_is_number = bool(_re.fullmatch(r'-?[\d.]+', after_token))

        # Last two tokens before the dash
        before_tokens = _re.split(r'\s+', before_text.strip()) if before_text.strip() else []
        prev_token  = before_tokens[-1] if before_tokens else ''
        prev2_token = before_tokens[-2] if len(before_tokens) >= 2 else ''

        prev_is_number = bool(_re.fullmatch(r'[\d.]+[\)\]]*', prev_token.strip('()')))
        prev2_is_op    = (prev2_token in OPERATORS or
                          any(c in OPERATORS for c in prev2_token) or
                          bool(_re.fullmatch(r'[\d.]+', prev2_token)))

        is_arithmetic = after_is_number or (prev_is_number and prev2_is_op)

        if after_char.isalpha() and not is_arithmetic:
            result.append(before_text)
            last = m.end()

    result.append(text[last:])
    if len(result) == 1:
        return text
    lines = [result[0]] + ['- ' + p for p in result[1:]]
    return '\n'.join(lines)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _jama_sheet(wb: Workbook, result: "MergeResult"):
    ws = wb.active
    ws.title = "Jama_Requirements"
    ws.sheet_view.showGridLines = False

    cols = [
        "ID",
        "Name",
        "Description",
        "VerificationMethod",
        "RequirementType",
        "Function (Source)",
    ]
    _banner(ws, "Jama High-Level Requirements — Import Sheet", len(cols))
    _headers(ws, cols)

    for i, hlr in enumerate(result.hlrs, start=3):
        alt  = (i % 2 == 0)
        _cell(ws, i, 1, hlr.id, alt)
        _cell(ws, i, 2, hlr.name, alt)
        _cell(ws, i, 3, _fmt_description(hlr.description), alt)
        _cell(ws, i, 4, hlr.verification_method, alt)
        _cell(ws, i, 5, hlr.requirement_type, alt)
        _cell(ws, i, 6, hlr.function_name, alt)
        ws.row_dimensions[i].height = 65

    _widths(ws, {1: 24, 2: 32, 3: 68, 4: 22, 5: 24, 6: 22, 7: 26, 8: 18})
    ws.freeze_panes = "A3"

def _summary_sheet(wb: Workbook, result: "MergeResult"):
    ws = wb.create_sheet("Function_Summary")
    ws.sheet_view.showGridLines = False

    cols = ["Function Name", "LLR Count", "HLRs Generated", "HLR IDs"]
    _banner(ws, "Function Summary", len(cols))
    _headers(ws, cols)

    hlrs_by_func: dict[str, list] = defaultdict(list)
    for hlr in result.hlrs:
        hlrs_by_func[hlr.function_name].append(hlr)

    for i, group in enumerate(result._groups, start=3):
        hlrs  = hlrs_by_func.get(group.function_name, [])
        alt   = (i % 2 == 0)
        warn  = len(hlrs) == 0
        # Support all group shapes across merger modes:
        #   merger.py          -> group.requirements (list of LLR strings)
        #   merger_rewrite.py  -> group.draft        (single combined string)
        if hasattr(group, "requirements"):
            count = len(group.requirements)
        elif hasattr(group, "drafts"):
            count = len(group.drafts)
        else:
            count = 1  # single draft string — one input per function
        _cell(ws, i, 1, group.function_name, alt, warn)
        _cell(ws, i, 2, str(count), alt, warn)
        _cell(ws, i, 3, str(len(hlrs)) if hlrs else "0 — skipped (empty)", alt, warn)
        _cell(ws, i, 4, ", ".join(h.id for h in hlrs) if hlrs else "—", alt, warn)
        ws.row_dimensions[i].height = 22

    _widths(ws, {1: 32, 2: 12, 3: 22, 4: 36})
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_to_excel(
    result: "MergeResult",
    output_path: str,
    filename
) -> None:
    """
    Write the Jama Excel workbook to output_path.

    Parameters
    ----------
    result      : MergeResult produced by merger.py or merger_rewrite.py
    output_path : Destination .xlsx file path
    module_name : Title for the root node of the hierarchy sheet
                  (passed via --module-name CLI argument, default shown above)
    """
    filename = filename.split('_')[0]
    module_name: str = f"Software Requirements for {filename} Module"
    from hierarchy_sheet import build_hierarchy_sheet
    wb = Workbook()
    _jama_sheet(wb, result)
    build_hierarchy_sheet(wb, result, module_name=module_name)
    wb.save(output_path)
